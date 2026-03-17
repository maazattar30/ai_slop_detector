"""
app.py — Gradio interface
Password protected AI video detector
"""

import gradio as gr
import json
import os
import base64
import tempfile
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import PASSWORD, BUCKETS
from pipeline import run_pipeline


# ─────────────────────────────────────────────
# PASSWORD CHECK
# ─────────────────────────────────────────────

def check_password(password: str) -> bool:
    if not PASSWORD:
        return True
    return password.strip() == PASSWORD.strip()


# ─────────────────────────────────────────────
# EXCEL REPORT GENERATOR
# ─────────────────────────────────────────────

def _generate_excel(result: dict) -> str:
    """Generate Excel report and return file path."""
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Styles
    header_font    = Font(bold=True, color="FFFFFF", size=12)
    header_fill    = PatternFill("solid", start_color="1F4E79")
    bucket_colors  = {"1": "16A34A", "2": "2563EB", "3": "DC2626"}
    bucket_color   = bucket_colors.get(str(result.get("bucket", 2)), "2563EB")
    verdict_fill   = PatternFill("solid", start_color=bucket_color)
    label_font     = Font(bold=True, color="1F2937", size=11)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def set_header(cell, value):
        cell.value = value
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    def set_label(cell, value):
        cell.value = value
        cell.font  = label_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    def set_value(cell, value):
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # Title row
    ws.merge_cells("A1:D1")
    ws["A1"].value     = "🎬 AI Slop Detector — Analysis Report"
    ws["A1"].font      = Font(bold=True, size=16, color="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Verdict row
    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"Verdict: Bucket {result.get('bucket')} — {result.get('bucket_name')} | Score: {result.get('final_score')}/100 | Confidence: {result.get('confidence', 0):.0%}"
    ws["A2"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A2"].fill      = verdict_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    # Blank row
    ws.row_dimensions[3].height = 8

    # Video info header
    ws.merge_cells("A4:D4")
    set_header(ws["A4"], "VIDEO INFORMATION")
    ws.row_dimensions[4].height = 22

    info_rows = [
        ("URL",             result.get("url", "")),
        ("Title",           result.get("title", "")),
        ("Channel",         result.get("video_info", {}).get("channel", "")),
        ("Duration (s)",    result.get("video_info", {}).get("duration", "")),
        ("Upload Date",     result.get("video_info", {}).get("upload_date", "")),
        ("Views",           result.get("video_info", {}).get("view_count", "")),
        ("Likes",           result.get("video_info", {}).get("like_count", "")),
        ("Subscribers",     result.get("video_info", {}).get("subscriber_count", "")),
    ]
    for i, (label, value) in enumerate(info_rows, start=5):
        set_label(ws.cell(row=i, column=1), label)
        ws.merge_cells(f"B{i}:D{i}")
        set_value(ws.cell(row=i, column=2), value)
        ws.row_dimensions[i].height = 18

    # Blank row
    row = 5 + len(info_rows) + 1

    # Scores header
    ws.merge_cells(f"A{row}:D{row}")
    set_header(ws.cell(row=row, column=1), "SCORES")
    ws.row_dimensions[row].height = 22
    row += 1

    score_rows = [
        ("Module Score",  result.get("module_score", 0)),
        ("LLM Score",     result.get("llm_score", 0)),
        ("Final Score",   result.get("final_score", 0)),
        ("Confidence",    f"{result.get('confidence', 0):.0%}"),
        ("Bucket",        f"{result.get('bucket')} — {result.get('bucket_name')}"),
    ]
    for label, value in score_rows:
        set_label(ws.cell(row=row, column=1), label)
        ws.merge_cells(f"B{row}:D{row}")
        set_value(ws.cell(row=row, column=2), value)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # LLM explanation header
    ws.merge_cells(f"A{row}:D{row}")
    set_header(ws.cell(row=row, column=1), "LLM EXPLANATION")
    ws.row_dimensions[row].height = 22
    row += 1

    llm = result.get("llm_result", {})
    ws.merge_cells(f"A{row}:D{row+2}")
    cell = ws.cell(row=row, column=1)
    cell.value     = llm.get("explanation", "")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border    = thin_border
    ws.row_dimensions[row].height   = 60
    row += 3

    # Red flags
    ws.merge_cells(f"A{row}:D{row}")
    set_header(ws.cell(row=row, column=1), "RED FLAGS")
    ws.row_dimensions[row].height = 22
    row += 1

    red_flags = llm.get("red_flags", [])
    if red_flags:
        for flag in red_flags:
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1)
            cell.value  = f"⚠ {flag}"
            cell.font   = Font(color="DC2626")
            cell.border = thin_border
            ws.row_dimensions[row].height = 18
            row += 1
    else:
        ws.merge_cells(f"A{row}:D{row}")
        set_value(ws.cell(row=row, column=1), "None detected")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ── Sheet 2: Signal Breakdown ─────────────────────────
    ws2 = wb.create_sheet("Signal Breakdown")

    headers = ["Feature", "Value", "Signal", "AI Probability", "Ranges", "Definition"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col)
        set_header(cell, h)
        ws2.row_dimensions[1].height = 22

    signal_fill_map = {
        "STRONG AI":     "FEE2E2",
        "MODERATE AI":   "FEF3C7",
        "NEUTRAL":       "F1F5F9",
        "MODERATE REAL": "DBEAFE",
        "STRONG REAL":   "DCFCE7",
    }

    for i, row_data in enumerate(result.get("signal_rows", []), start=2):
        label     = row_data.get("label", "NEUTRAL")
        fill_color = signal_fill_map.get(label, "F1F5F9")
        row_fill   = PatternFill("solid", start_color=fill_color)

        values = [
            row_data.get("feature", ""),
            row_data.get("value", ""),
            label,
            f"{row_data.get('ai_prob', 0):.2f}",
            row_data.get("ranges", ""),
            row_data.get("definition", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell        = ws2.cell(row=i, column=col, value=val)
            cell.fill   = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.row_dimensions[i].height = 22

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 35
    ws2.column_dimensions["F"].width = 45

    # Save
    safe_title = "".join(
        c for c in result.get("title", "report")[:30]
        if c.isalnum() or c in " _-"
    ).strip() or "report"
    out_path = os.path.join(tempfile.gettempdir(), f"ai_slop_{safe_title}.xlsx")
    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────
# MAIN ANALYSIS FUNCTION
# ─────────────────────────────────────────────

def analyze(
    url:         str,
    title:       str,
    description: str,
    password:    str,
    progress=gr.Progress(track_tqdm=True)
) -> tuple:
    """
    Returns (status_text, result_json, report_html, excel_file, frames_gallery)
    """
    if not check_password(password):
        return (
            "❌ Incorrect password", "{}", 
            "<p style='color:red;padding:20px'>Access denied</p>",
            None, []
        )

    url = url.strip()
    if not url:
        return (
            "❌ Please enter a YouTube URL", "{}",
            "<p style='color:red;padding:20px'>No URL provided</p>",
            None, []
        )

    if "youtube" not in url and "youtu.be" not in url:
        return (
            "❌ Please enter a valid YouTube URL", "{}",
            "<p style='color:red;padding:20px'>Invalid URL</p>",
            None, []
        )

    def progress_cb(pct, msg):
        progress(pct, desc=msg)

    try:
        progress(0, desc="Starting pipeline...")
        result = run_pipeline(
            url         = url,
            title       = title,
            description = description,
            progress_cb = progress_cb,
        )

        if result.get("error"):
            return (
                f"❌ Error: {result['error']}",
                json.dumps(result, indent=2),
                f"<p style='color:red;padding:20px'>Error: {result['error']}</p>",
                None, []
            )

        html        = _build_report_html(result)
        excel_path  = _generate_excel(result)
        frames      = _get_top_frames(result)

        status = (
            f"✅ Done — {result['bucket_name']} "
            f"(Score: {result['final_score']}/100)"
        )

        return (
            status,
            json.dumps(result, indent=2, default=str),
            html,
            excel_path,
            frames
        )

    except Exception as e:
        return (
            f"❌ Unexpected error: {str(e)}", "{}",
            f"<p style='color:red;padding:20px'>{str(e)}</p>",
            None, []
        )


def _get_top_frames(result: dict) -> list:
    """Return top 3 LLM frame paths if bucket is AI Generated (3)."""
    if result.get("bucket") != 3:
        return []
    frames = result.get("llm_frames", [])
    return [f for f in frames[:3] if os.path.exists(f)]


# ─────────────────────────────────────────────
# HTML REPORT BUILDER
# ─────────────────────────────────────────────

def _build_report_html(result: dict) -> str:
    bucket      = result["bucket"]
    b_info      = BUCKETS[bucket]
    final_score = result["final_score"]
    mod_score   = result["module_score"]
    llm_score   = result["llm_score"]
    llm         = result.get("llm_result", {})
    confidence  = result.get("confidence", 0.5)
    color       = b_info["color"]
    emoji       = b_info["emoji"]
    name        = b_info["name"]

    def score_bar(score, bar_color):
        pct = min(max(score, 0), 100)
        return (
            f"<div style='background:#eee;border-radius:6px;height:10px;margin:6px 0'>"
            f"<div style='background:{bar_color};width:{pct}%;height:10px;border-radius:6px'></div></div>"
        )

    def pill(label):
        colors = {
            "STRONG AI":     ("background:#FEE2E2", "color:#DC2626"),
            "MODERATE AI":   ("background:#FEF3C7", "color:#D97706"),
            "NEUTRAL":       ("background:#F1F5F9", "color:#64748B"),
            "MODERATE REAL": ("background:#DBEAFE", "color:#2563EB"),
            "STRONG REAL":   ("background:#DCFCE7", "color:#16A34A"),
        }
        bg, fg = colors.get(label, ("background:#F1F5F9", "color:#94A3B8"))
        return (
            f"<span style='{bg};{fg};padding:2px 8px;border-radius:99px;"
            f"font-size:11px;font-weight:700'>{label}</span>"
        )

    sig_rows   = result.get("signal_rows", [])
    table_rows = ""
    for row in sig_rows:
        table_rows += (
            f"<tr style='border-bottom:1px solid #F1F5F9'>"
            f"<td style='padding:8px 12px;font-weight:600;color:#2C3E50'>{row['feature']}</td>"
            f"<td style='padding:8px 12px;font-family:monospace;color:#17202A'>{row['value']}</td>"
            f"<td style='padding:8px 12px'>{pill(row['label'])}</td>"
            f"<td style='padding:8px 12px;color:#64748B;font-size:12px'>{row.get('ranges','')}</td>"
            f"<td style='padding:8px 12px;color:#64748B;font-size:12px'>{row.get('definition','')}</td>"
            f"</tr>"
        )

    def ul(items, item_color="#334155"):
        if not items:
            return "<li style='color:#94A3B8'>None detected</li>"
        return "".join(f"<li style='margin:4px 0;color:{item_color}'>{i}</li>" for i in items)

    prim_html  = ul(llm.get("primary_signals", []))
    vis_html   = ul(llm.get("visual_evidence", []))
    flags_html = ul(llm.get("red_flags", []), "#DC2626")

    override_html = ""
    if llm.get("module_override"):
        override_html = (
            f"<div style='background:#FEF9E7;border-left:4px solid #F39C12;"
            f"padding:12px;margin:12px 0;border-radius:4px'>"
            f"<b>⚠️ LLM override:</b> {llm.get('override_reason', '')}</div>"
        )

    html = f"""
<div style='font-family:Arial,sans-serif;max-width:1000px;margin:0 auto;padding:20px'>
  <div style='background:linear-gradient(135deg,#0F172A,#1E3A5F);color:white;
              padding:22px;border-radius:12px;margin-bottom:18px'>
    <h2 style='margin:0;font-size:22px'>🎬 AI Slop Detector</h2>
    <p style='margin:4px 0 0 0;opacity:0.8;font-size:13px'>{result.get('title','')[:80]}</p>
  </div>
  <div style='background:{color}15;border:2px solid {color};border-radius:12px;
              padding:24px;margin-bottom:18px;text-align:center'>
    <div style='font-size:48px'>{emoji}</div>
    <div style='font-size:28px;font-weight:900;color:{color};margin:6px 0'>
      Bucket {bucket} — {name}
    </div>
    <div style='color:#566573'>
      AI Score: <b style='color:{color}'>{final_score}</b>/100
      &nbsp;|&nbsp;
      Confidence: <b style='color:{color}'>{confidence:.0%}</b>
    </div>
  </div>
  <div style='display:flex;gap:14px;margin-bottom:18px'>
    <div style='flex:1;background:white;border:1px solid #E2E8F0;border-radius:8px;padding:16px'>
      <div style='font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em'>Module Score</div>
      <div style='font-size:30px;font-weight:700;color:#2563EB'>{mod_score:.1f}</div>
      {score_bar(mod_score, "#2563EB")}
      <div style='font-size:11px;color:#94A3B8'>12 signals, Cohen d-weighted</div>
    </div>
    <div style='flex:1;background:white;border:1px solid #E2E8F0;border-radius:8px;padding:16px'>
      <div style='font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em'>LLM Score</div>
      <div style='font-size:30px;font-weight:700;color:#7C3AED'>{llm_score}</div>
      {score_bar(llm_score, "#7C3AED")}
      <div style='font-size:11px;color:#94A3B8'>Llama 4 Scout visual reasoning</div>
    </div>
    <div style='flex:1;background:{color}10;border:2px solid {color};border-radius:8px;padding:16px'>
      <div style='font-size:11px;color:#94A3B8;text-transform:uppercase;letter-spacing:0.1em'>Final Score</div>
      <div style='font-size:30px;font-weight:700;color:{color}'>{final_score}</div>
      {score_bar(final_score, color)}
      <div style='font-size:11px;color:#94A3B8'>55% module + 45% LLM</div>
    </div>
  </div>
  {override_html}
  <div style='background:#EBF5FB;border-left:4px solid #2563EB;padding:16px;
              border-radius:4px;margin-bottom:18px'>
    <b style='color:#1F4E79'>📋 Explanation:</b><br>
    <p style='margin:8px 0 0 0;color:#2C3E50;line-height:1.7'>
      {llm.get('explanation', 'No explanation available.')}
    </p>
  </div>
  <div style='display:flex;gap:14px;margin-bottom:18px'>
    <div style='flex:1;background:white;border:1px solid #E2E8F0;border-radius:8px;padding:14px'>
      <b style='color:#1F4E79'>🎯 Primary Signals</b>
      <ul style='margin:8px 0 0 0;padding-left:18px'>{prim_html}</ul>
    </div>
    <div style='flex:1;background:white;border:1px solid #E2E8F0;border-radius:8px;padding:14px'>
      <b style='color:#1F4E79'>👁 Visual Evidence</b>
      <ul style='margin:8px 0 0 0;padding-left:18px'>{vis_html}</ul>
    </div>
    <div style='flex:1;background:white;border:1px solid #E2E8F0;border-radius:8px;padding:14px'>
      <b style='color:#DC2626'>🚩 AI Red Flags</b>
      <ul style='margin:8px 0 0 0;padding-left:18px'>{flags_html}</ul>
    </div>
  </div>
  <div style='background:white;border:1px solid #E2E8F0;border-radius:8px;margin-bottom:18px'>
    <div style='padding:14px;border-bottom:1px solid #F1F5F9'>
      <b style='color:#1F4E79;font-size:15px'>📊 Signal Breakdown</b>
    </div>
    <div style='overflow-x:auto'>
      <table style='width:100%;border-collapse:collapse;font-size:13px'>
        <thead>
          <tr style='background:#1F4E79;color:white'>
            <th style='padding:10px 12px;text-align:left'>Feature</th>
            <th style='padding:10px 12px;text-align:left'>Value</th>
            <th style='padding:10px 12px;text-align:left'>Signal</th>
            <th style='padding:10px 12px;text-align:left'>Ranges</th>
            <th style='padding:10px 12px;text-align:left'>What it measures</th>
          </tr>
        </thead>
        <tbody>{table_rows}</tbody>
      </table>
    </div>
  </div>
  <div style='color:#94A3B8;font-size:11px;text-align:center;padding:8px'>
    AI Slop Detector V1 &nbsp;|&nbsp; Signals: 10 modules &nbsp;|&nbsp; LLM: Llama 4 Scout (Groq)
  </div>
</div>"""
    return html


# ─────────────────────────────────────────────
# GRADIO UI
# ─────────────────────────────────────────────

with gr.Blocks(title="AI Slop Detector", theme=gr.themes.Soft()) as demo:

    gr.Markdown("""
    # 🎬 AI Slop Detector
    Paste a YouTube URL to classify it as **Human Only**, **Human + AI Tools**, or **AI Generated**
    using 10 signal modules + Llama 4 Scout visual reasoning.
    """)

    with gr.Row():
        with gr.Column(scale=2):
            url_input   = gr.Textbox(label="YouTube URL *", placeholder="https://www.youtube.com/watch?v=...")
            title_input = gr.Textbox(label="Video Title (optional — auto-fetched)", placeholder="e.g. My Travel Vlog")
            desc_input  = gr.Textbox(label="Description (optional)", placeholder="Paste description for text signal analysis...", lines=3)
            pass_input  = gr.Textbox(label="Password", type="password", placeholder="Enter access password")
            run_btn     = gr.Button("▶  Analyze Video", variant="primary", size="lg")

        with gr.Column(scale=1):
            gr.Markdown("""
            ### Bucket Reference

            🟢 **B1 — Human Only**
            Real camera, real people, no AI

            🔵 **B2 — Human + AI Tools**
            Human-made with AI assistance.
            Also used when uncertain.

            🔴 **B3 — AI Generated**
            Text-to-video, deepfake, TTS, AI avatar

            ---
            ⏱ **Processing time:** ~2-4 min

            🔬 **Signals:** 10 modules

            🤖 **LLM:** Llama 4 Scout (Groq)
            """)

    status_out = gr.Textbox(label="Status", interactive=False)

    with gr.Tabs():
        with gr.Tab("📊 Report Card"):
            report_html = gr.HTML()

        with gr.Tab("🖼 AI Sample Frames"):
            gr.Markdown("### Top 3 AI Sample Frames *(only shown for AI Generated videos)*")
            frames_gallery = gr.Gallery(
                label   = "AI Evidence Frames",
                columns = 3,
                height  = 400,
                show_label = False,
            )

        with gr.Tab("🔧 Raw JSON"):
            json_out = gr.Code(language="json")

    # Excel download button
    with gr.Row():
        excel_out = gr.File(label="📥 Download Excel Report", visible=True)

    run_btn.click(
        fn      = analyze,
        inputs  = [url_input, title_input, desc_input, pass_input],
        outputs = [status_out, json_out, report_html, excel_out, frames_gallery]
    )

if __name__ == "__main__":
  port = int(os.environ.get("PORT", 7860))

  demo.queue(
      max_size=5,
      default_concurrency_limit=1
  )

  demo.launch(
      server_name="0.0.0.0",
      server_port=port,
      share=False,
      max_threads=1
  )